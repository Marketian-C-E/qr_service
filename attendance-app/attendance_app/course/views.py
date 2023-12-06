import os

from django.shortcuts import render, redirect, get_object_or_404, get_list_or_404

from course.forms import CourseForm, ClassAttendInForm
from survey.forms import SurveyForm

from course.models import Course, ClassAttend
from user.models import Division
from user.models import Student
from survey.models import SurveyReply,Survey
from django.db.models import Q

from django.contrib.auth.decorators import login_required, user_passes_test

# 엑셀 다운로드
import openpyxl
from openpyxl.utils import get_column_letter
from django.conf import settings

from datetime import datetime, timedelta
from django.http import HttpResponse



# Create your views here.

# QR코드 스캐너 & 스캔후 데이터 받아서 출결 찍기
def QRScanner_in(request, pk):
    course = get_object_or_404(Course, pk=pk)

    if request.method == 'POST' and  'data' in request.POST and request.POST.get('data') != "":
        qr_data = request.POST.get('data')
        
        try:
            student = Student.objects.get(name=qr_data)
            # user Id값이 data에 들어감
            print("url 데이터 : " + student.name)

            # 현재 시간( 년도-월-일-시각-분)
            now = datetime.now()
            time_now = now.strftime('%H:%M:%S')
                

            context = {
                'course' : course,
                'check_time' : time_now,
                'student' : student,
                'status' : True,    # 입실 -> True, 퇴실 -> False
            }
            
            # 출석체크 페이지로 이동
            return render(request, 'attendance/attendance_check.html', context)
        except Student.DoesNotExist:
            print("유효한 QR코드가 입력값으로 들어오지 않았습니다.")
            context = {
                'course' : course,
                'qr_error' : True,
                'error_message' : "유효한 QR코드가 아닙니다. 등록된 QR코드를 보여주세요"
            }
            return render(request, 'attendance/QRScanner_in.html', context)
        
    else :
        context = {
            'course' : course,
        }
        return render(request, 'attendance/QRScanner_in.html', context)


def QRScanner_out(request, pk):
    course = get_object_or_404(Course, pk=pk)

    if request.method == 'POST' and  'data' in request.POST and request.POST.get('data') != "":
        qr_data = request.POST.get('data')
        
        try:
            student = Student.objects.get(name=qr_data)
            # user Id값이 data에 들어감
            print("url 데이터 : " + student.name)

            # 현재 시간( 년도-월-일-시각-분)
            now = datetime.now()
            time_now = str(now.strftime('%Y-%m-%d %H:%M'))
                

            context = {
                'course' : course,
                'check_time' : time_now,
                'student' : student,
                'status' : False, # 입실 -> True, 퇴실 -> False
            }
            
            # 출석체크 페이지로 이동
            return render(request, 'attendance/attendance_check.html', context)
        except Student.DoesNotExist:
            print("유효한 QR코드가 입력값으로 들어오지 않았습니다.")
            context = {
                'course' : course,
                'qr_error' : True,
                'error_message' : "유효한 QR코드가 아닙니다. 등록된 QR코드를 보여주세요"
            }
            return render(request, 'attendance/QRScanner_out.html', context)
        
    else :
        context = {
            'course' : course,
        }

        return render(request, 'attendance/QRScanner_out.html', context)

### 입실,퇴실 체크 모듈에 데이터 넣기

#입실용
def attendance_check_in(request):
    if request.method == 'POST':
        print("post입력 확인")
        form = ClassAttendInForm(request.POST)
        class_attend = form.save(commit=False)
        #  출석 데이터 조회
        attend_data = ClassAttend.objects.filter(Q(course_id=class_attend.course_id) & Q(student_id=class_attend.student_id))
        
        if len(attend_data) == 0:
            # 데이터가 0개인 경우
            if form.is_valid():
                # 정상 입실 확인 / ClassAttend 객체 생성
                print("form 유효성 성공")
                classAttend = form.save()
                return redirect('course:attendance_check_in_success', pk=classAttend.pk)  # 식별자(pk)를 URL에 포함시켜 리디렉션
            else:
                errors = form.errors.as_text()
                print(errors)  # 에러 메시지 출력 또는 원하는 동작 수행
                return render(request, 'attendance/attendance_error.html', {'form': form})
        else :
            # 데이터가 1개 이상인 경우 (출입을 이미 한 경우)
            course = class_attend.course_id
            student = class_attend.student_id
            error_message = student.name + "님 이미 입실등록을 하셧습니다."
            context = {
                'course': course, 
                'error_message': error_message,
                'qr_error' : True,
            }
            return render(request, 'attendance/QRScanner_in.html', context)
    else:
        return render(request, 'attendance/attendance_error.html')

# 출석 성공 페이지
def attendance_check_in_success(request, pk):

    classAttend = ClassAttend.objects.get(pk=pk)  # 식별자(pk)를 사용하여 클래스 인스턴스 조회

    # 시간값 비교를 위한 datetime 객체로 변환. 시간값만 비교
    course_late_at = datetime.combine(datetime.min,classAttend.course_id.start_at) + timedelta(minutes=15)
    print(course_late_at)
    student_start_at = datetime.combine(datetime.min,classAttend.start_at)
    print(student_start_at)

    # 시작시간보다 15분 뒤에 입실을 할시 (지각 처리)
    if student_start_at > course_late_at :
        late_check = False
    # 시작 시간 15분 이내에 입실을 할시 (정상 출석)
    else :
        late_check = True

    course = classAttend.course_id
    student = classAttend.student_id

    # 현재 듣고있는 강의 필드에 추가
    student.current_course_name = course.course_name
    student.save()

    context =  {
        'classAttend': classAttend, 
        'course': course, 
        'student': student,
        'late_check' : late_check,
    }
    return render(request, 'attendance/attendance_check_in_success.html', context)

################################
# (미사용)퇴실용
def attendance_check_out(request):
    if request.method == 'POST':
        course_id = request.POST.get('course_id')
        student_id = request.POST.get('student_id')

        end_at = request.POST.get('end_at')
        time_string = end_at
        parsed_time = datetime.strptime(time_string, "%Y-%m-%d %H:%M")
        formatted_time = parsed_time.strftime("%H:%M")
        print(formatted_time)
        

        # 지각을해서 입실처리는 하지않고, 퇴실처리만 시도한 경우
        try : 
            classAttend = ClassAttend.objects.get(Q(course_id=course_id) & Q(student_id=student_id))
            print(classAttend)
        except :
            student = Student.objects.get(id = student_id)
            error_message = student.name + "님 입실 처리를 하셔야 퇴실 처리를 하실수 있습니다!! (지각시 관리자에게 수동 출석을 요청하세요)"
            context = {
                'course' : Course.objects.get(id = course_id),
                'error_message' : error_message,
                'qr_error' : True,
            }
            return render(request, 'attendance/QRScanner_out.html', context)            
        
        # 설문지 제출했는지 확인
        # survey_reply = SurveyReply.objects.get(Q(course_id=course_id) & Q(student_id=student_id))
        # if survey_reply.submit_survey !=True:
        #    ...
        
        ### 퇴실 예외처리 부분
        # 강의평가를 제출했는지 확인 
        survey = Survey.objects.get(course_id=classAttend.course_id)
        try :
            survey_reply = SurveyReply.objects.get(student_id=classAttend.student_id, survey_id=survey)
            # 2차 검증 (강의평가를 제출하지 않았으면)
            if survey_reply.submit_survey == False:
                course = classAttend.course_id
                student = classAttend.student_id
                error_message = student.name + "님 강의평가를 제출하지 않으셔서 퇴실처리 되지않았습니다. (강의평가 제출 후 퇴실처리 바랍니다.)"
                context = {
                    'course' : course,
                    'error_message' : error_message,
                    'qr_error' : True,
                }
                return render(request, 'attendance/QRScanner_out.html', context)            
            
        except SurveyReply.DoesNotExist : 
            # 강의평가를 제출하지 않았으면 
            course = classAttend.course_id
            student = classAttend.student_id
            error_message = student.name + "님 강의평가를 제출하지 않으셔서 퇴실처리 되지않았습니다. (강의평가 제출 후 퇴실처리 바랍니다.)"
            context = {
                'course' : course,
                'error_message' : error_message,
                'qr_error' : True,
            }
            return render(request, 'attendance/QRScanner_out.html', context)            
        
        # classAttend.attend_state(퇴실 처리를 이미 했으면)
        if classAttend.attend_state == 2:
            course = classAttend.course_id
            student = classAttend.student_id
            error_message = student.name + "님 이미 정상 퇴실처리 되셧습니다."
            context = {
                'course' : course,
                'error_message' : error_message,
                'qr_error' : True,
            }
            return render(request, 'attendance/QRScanner_out.html', context)            
        

        ### 정상 퇴실 처리 부분 

        # 시간값 비교를 위한 datetime 객체로 변환. 시간값만 비교
        course_late_at = datetime.combine(datetime.min,classAttend.course_id.start_at) + timedelta(minutes=15)
        print(course_late_at)
        student_start_at = datetime.combine(datetime.min,classAttend.start_at)
        print(student_start_at)

        # 시작시간보다 15분 뒤에 입실을 할시 (지각 처리)
        if student_start_at > course_late_at :
            classAttend.end_at = formatted_time
            classAttend.attend_state = 1 # 지각 출석
            classAttend.save()
        # 시작 시간 15분 이내에 입실을 할시 (정상 출석)
        else :
            classAttend.end_at = formatted_time
            classAttend.attend_state = 2 # 정상 출석
            classAttend.save()
            
        course = classAttend.course_id
        student = classAttend.student_id

        # 퇴실까지 완료하면 다시 리셋
        student.current_course_name = "수강중인 강의 없음"
        student.save()

        # 건네줄 cotext
        context = {
            'course' : course,
            'student' : student,
            'classAttend' : classAttend,
        }
        return render(request,'attendance/attendance_check_out.html', context)
    else:
       return render(request, 'attendance/attendance_error.html')
##############################################################



# 강좌 리스트
# Create your views here.
@user_passes_test(lambda u: u.is_staff, login_url='/') # 권한 없으면 홈으로
def course_list(request, pk):
    division = Division.objects.get(pk=pk)
    course = Course.objects.filter(division_name_id=pk)
    
    context = {
        'course': course,
        'division': division,
        
        }
    return render(request, 'course/course_list.html', context)


# 강좌 추가
@user_passes_test(lambda u: u.is_staff, login_url='/') # 권한 없으면 홈으로
def create_course(request):
    if request.method == 'POST':
        course_form = CourseForm(request.POST)
        survey_form = SurveyForm(request.POST)
        
        if course_form.is_valid() and survey_form.is_valid():
            # 코스 저장
            course = course_form.save()
            survey = survey_form.save(commit=False)
            
            survey.course_id_id = course.id
            survey_form.save()
  
            return redirect('user:division_list')  # 적절한 URL로 리다이렉트


    else:
        course_form = CourseForm()
        survey_form = SurveyForm()

    context = {
        'course_form': course_form,
        'survey_form': survey_form,
    }
    
    return render(request, 'course/create_course.html', context)


# 강좌 상세 페이지
@user_passes_test(lambda u: u.is_staff, login_url='/') # 권한 없으면 홈으로
def course_detail(request, pk):
    course = get_object_or_404(Course, pk=pk)
    context = {'course': course}
    return render(request, 'course/course_detail.html', context)


# 강좌 수정
@user_passes_test(lambda u: u.is_staff, login_url='/') # 권한 없으면 홈으로
def edit_course(request, pk):
    course = get_object_or_404(Course, pk=pk)
    
    if request.method == 'POST':
        form = CourseForm(request.POST, instance=course)
        if form.is_valid():
            form.save()
            return redirect('course:course_detail', pk=course.pk)
        
    else:
        form = CourseForm(instance=course)
    context = {'form': form, 'course':course}
    return render(request, 'course/edit_course.html', context)


# 강좌 삭제
@user_passes_test(lambda u: u.is_staff, login_url='/') # 권한 없으면 홈으로
def delete_course(request, pk):
    course = get_object_or_404(Course, pk=pk)
    if request.method == 'POST':
        course.delete()
        return redirect('user:division_list')  # Replace 'home' with the appropriate URL name
    
    context = {'course': course}
    return render(request, 'course/delete_course.html', context)


# 출석 분반 선택
@user_passes_test(lambda u: u.is_staff, login_url='/') # 권한 없으면 홈으로
def attendance_divison_list(request):
    division = Division.objects.all()

    context = {
        'division': division
    }

    return render(request, 'attendance_board/attendance_division_list.html', context)


# 출석 강의 선택
@user_passes_test(lambda u: u.is_staff, login_url='/') # 권한 없으면 홈으로
def attendance_course_list(request, pk):
    # pk -> division_id
    course = Course.objects.filter(division_name_id=pk)

    context = {
        'course': course
    }

    return render(request, 'attendance_board/attendance_course_list.html', context)



# 강의별 출석 명단
@user_passes_test(lambda u: u.is_staff, login_url='/') # 권한 없으면 홈으로
def attendance_course_board(request, pk):
    # pk -> course_id
    course = Course.objects.get(pk=pk)

    division = Division.objects.get(pk=course.division_name_id)
    
    students = Student.objects.filter(division_id=division.pk).order_by('name')
    class_attends = ClassAttend.objects.filter(Q(student_id__division_id=division.pk) & Q(course_id_id=course))
    

    print(f'{class_attends = }')

    context = {
        'course': course,
        'class_attends': class_attends,
        'students': students,
        'division': division,
        
    }

    return render(request, 'attendance_board/attendance_course_board.html', context)



# 출석부에서 출결 변경시 처리
@user_passes_test(lambda u: u.is_staff, login_url='/') # 권한 없으면 홈으로
def student_attendance_update(request):
    if request.method == 'POST':
        # 폼으로 전송된 데이터 가져오기
        search_mode = request.POST.get('search_mode')
        print(search_mode)
        student_id = request.POST.get('student_id')
        print(student_id)
        course_id = request.POST.get('course_id')
        print(course_id)
        
        student = Student.objects.get(id = student_id)
        course = Course.objects.get(id = course_id)
        
        student_attend = None
        try:
            student_attend = ClassAttend.objects.get(Q(course_id=course) & Q(student_id=student))
        except:
            student_attend = ClassAttend(course_id= course, student_id=student)
            student_attend.save()
    
        print(student_attend)

        # 2 출석, 1 지각, 0 결석, nftf_in 비대면 입실, nftf_out 비대면 퇴실
        if search_mode == "2" :
            student_attend.attend_state = 2
            student_attend.save()
        elif search_mode =="1" :
            student_attend.attend_state = 1
            student_attend.save()
        elif search_mode == "0" :
            student_attend.attend_state = 0
            student_attend.save()
        elif search_mode == "nftf_in" :
            parsed_time = datetime.now()
            formatted_time = parsed_time.strftime("%H:%M")
            student_attend.start_at = formatted_time
            student_attend.save()
        elif search_mode == "nftf_out" :
            parsed_time = datetime.now()
            formatted_time = parsed_time.strftime("%H:%M")
            student_attend.end_at = formatted_time
            student_attend.save()

        
        # 첫 출석체크(객체생성) 일 경우에만 student.current_course_name 값 설정. 1개 일때만.
        
        # student 수강중인 강의 상태 추가하기
        student_survey_reply = SurveyReply.objects.filter(Q(survey_id_id__course_id_id=course.pk) & Q(student_id_id=student_id))

        if len(student_survey_reply) < 1:
            student.current_course_name = f'{course.course_name}'
            student.save()
        
        
        return redirect("course:attendance_course_board", pk=course.pk)
    
    
    #return redirect
    
    
    
# 강의 출석현황 xlsx 다운로드
@user_passes_test(lambda u: u.is_staff, login_url='/') # 권한 없으면 홈으로
def download_attendance(request, pk):
    # pk -> course_id
    
    # 중간 query
    course = Course.objects.get(pk=pk)
    division = Division.objects.get(pk=course.division_name_id)
    
    students = Student.objects.filter(division_id=division.pk).order_by('name')
    class_attends = ClassAttend.objects.filter(Q(student_id__division_id=division.pk) & Q(course_id_id=course))  
    
    
    # 파일로 만들어 줄 최종 query
    
    
    print(f'{students = }')
    
    data = [['학생 이름', '입실시간', '설문조사 제출 시간 (퇴실)', '출석 인정', '분반 이름', '강의 이름']]
    for student in students:
        row = []  # 필요한 필드 값을 추출하여 리스트로 저장합니다.
        
        # 값 초기화
        student_name = student.name
        my_start_at = '---'
        my_end_at = '---'
        my_attend_state = '결석'
        division_name = division.name
        course_name = course.course_name
        
    
        for class_attend in class_attends:
            print(f'{class_attend.attend_state = }')
            print(f'{class_attend.student_id_id = }')
            print(f'{student.id = }')
            
            
            
            # 대응하는 기록 확인
            if class_attend.student_id_id == student.id:
                # 입실 기록 확인
                if class_attend.start_at != '00:00:00':
                    my_start_at = class_attend.start_at
                
    
                # 퇴실 기록 확인
                if class_attend.end_at != '00:00:00':
                    my_end_at = class_attend.end_at
                
                
                
                # 출석 인정 상태 표기
                if class_attend.attend_state == 0:
                    my_attend_state = '결석'
                    
                elif class_attend.attend_state == 1:
                    my_attend_state = '지각'
                
                elif class_attend.attend_state == 2:
                    my_attend_state = '출석'

        
        
        # 한 행 데이터 설정
        row.append(student_name)
        row.append(my_start_at)
        row.append(my_end_at)
        row.append(my_attend_state)
        row.append(division_name)
        row.append(course_name)
                
        # excel 데이터 한 행 추가
        data.append(row)
        

    ## 엑셀 파일 생성
    # 파일 이름
    excel_file = f'attendance_excel/출석부_강의_{course.course_name}_분반_{division.name}.xlsx'
    
    # 최종 경로
    excel_path = os.path.join(settings.MEDIA_ROOT, excel_file)
    
    # 만약 경로에 파일이 있으면 지우기, 이 부분은 없애도됨.
    # 혹시 모를 WINERROR 방지
    if os.path.exists(excel_path):
        os.remove(excel_path)
    
    
    # 엑셀 파일 열기
    wb = openpyxl.Workbook()
    sheet = wb.active

    for row_num, row_data in enumerate(data, start=1):
        for col_num, value in enumerate(row_data, start=1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}{row_num}"] = value
    
    
    with open(excel_path, 'wb') as f:
        wb.save(f)

    print(os.path.basename(excel_file))
    if os.path.exists(excel_path):
        with open(excel_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{excel_file}"'
            return response

